# -*- coding: utf-8 -*-
"""
測試 ExcelRenderer 與相關輔助元件（v2.2.1：樣板為主 + Jinja2 模板）
"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))


try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


requires_openpyxl = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl 不可用")


# --------------------------------------------------------------- layout


class TestLayoutConfig:
    @requires_openpyxl
    def test_default_values(self):
        from md_word_renderer.renderer.excel_layout import LayoutConfig
        cfg = LayoutConfig()
        assert cfg.default_template.endswith(".xlsx")
        assert cfg.header_sheet.enabled is True
        assert cfg.extra_columns == ["source_field", "path", "depth"]
        assert cfg.image.max_width_px == 480
        # v2.2.1 新增
        assert cfg.template_engine.enabled is True
        assert cfg.template_engine.auto_flatten_lists is True
        assert cfg.template_engine.missing_variable == "silent"

    @requires_openpyxl
    def test_from_mapping(self):
        from md_word_renderer.renderer.excel_layout import LayoutConfig
        data = {
            "default_template": "custom.xlsx",
            "list_sheet_naming": "{key}_items",
            "header_sheet": {"enabled": False, "name": "Header"},
            "image": {"max_width_px": 200, "max_height_px": 150},
            "auto_fit_columns": False,
            "extra_columns": ["a", "b"],
            "template_engine": {
                "enabled": True,
                "auto_flatten_lists": False,
                "missing_variable": "keep",
            },
        }
        cfg = LayoutConfig.from_mapping(data)
        assert cfg.default_template == "custom.xlsx"
        assert cfg.list_sheet_naming == "{key}_items"
        assert cfg.header_sheet.enabled is False
        assert cfg.header_sheet.name == "Header"
        assert cfg.image.max_width_px == 200
        assert cfg.auto_fit_columns is False
        assert cfg.extra_columns == ["a", "b"]
        assert cfg.template_engine.auto_flatten_lists is False
        assert cfg.template_engine.missing_variable == "keep"

    @requires_openpyxl
    def test_from_mapping_handles_missing(self):
        from md_word_renderer.renderer.excel_layout import LayoutConfig
        cfg = LayoutConfig.from_mapping({})
        assert cfg.image.max_width_px == 480
        assert cfg.header_sheet.enabled is True

    @requires_openpyxl
    def test_template_engine_invalid_mode_raises(self):
        from md_word_renderer.renderer.excel_layout import TemplateEngineConfig
        with pytest.raises(ValueError):
            TemplateEngineConfig(missing_variable="invalid_mode")


class TestSanitizeSheetName:
    @requires_openpyxl
    def test_strip_forbidden_chars(self):
        from md_word_renderer.renderer.excel_layout import sanitize_sheet_name
        assert sanitize_sheet_name("a/b\\c?d*e[f]g:h") == "a_b_c_d_e_f_g_h"

    @requires_openpyxl
    def test_empty_returns_fallback(self):
        from md_word_renderer.renderer.excel_layout import sanitize_sheet_name
        assert sanitize_sheet_name("") == "Sheet"
        assert sanitize_sheet_name("///") == "Sheet"

    @requires_openpyxl
    def test_truncate_to_31_chars(self):
        from md_word_renderer.renderer.excel_layout import sanitize_sheet_name
        long = "a" * 100
        result = sanitize_sheet_name(long)
        assert len(result) == 31


# ---------------------------------------------------- sheet-name de-dupe


@requires_openpyxl
def test_resolve_list_sheet_name_avoids_collisions():
    from md_word_renderer.renderer.excel_layout import LayoutConfig
    cfg = LayoutConfig()
    used = {"項目"}
    name = cfg.resolve_list_sheet_name("項目", used)
    assert name not in used
    assert name.startswith("項目")


# ---------------------------------------------------- image handler


class TestExcelImageHandler:
    @requires_openpyxl
    def test_raises_when_image_missing(self, tmp_path):
        from md_word_renderer.renderer.excel_image_handler import (
            ExcelImageHandler,
            ExcelImageError,
        )
        h = ExcelImageHandler()
        wb = Workbook()
        ws = wb.active
        with pytest.raises(ExcelImageError):
            h.embed(ws, "A1", str(tmp_path / "missing.png"))


# ---------------------------------------------------- template engine (Phase 1)


class TestExcelTemplateEngine:
    @requires_openpyxl
    def test_substitute_basic_var(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        result = engine.render_cell("{{name}}", {"name": "Alice"})
        assert result == "Alice"

    @requires_openpyxl
    def test_substitute_data_dict_special_chars(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        ctx = {"data": {"需求依據(INC/PBI)": "PBI123"}}
        result = engine.render_cell('{{data["需求依據(INC/PBI)"]}}', ctx)
        assert result == "PBI123"

    @requires_openpyxl
    def test_substitute_undefined_becomes_empty(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine(missing_variable="silent")
        assert engine.render_cell("{{missing}}", {}) == ""

    @requires_openpyxl
    def test_substitute_keep_mode_preserves_marker(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine(missing_variable="keep")
        assert engine.render_cell("{{missing}}", {}) == "{{missing}}"

    @requires_openpyxl
    def test_substitute_error_format_mode(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine(
            missing_variable="error_format",
            error_format="[ERROR: {var}]",
        )
        result = engine.render_cell("{{missing}}", {})
        assert "ERROR" in result
        assert "missing" in result

    @requires_openpyxl
    def test_engine_disabled_keeps_placeholders(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine(enabled=False)
        assert engine.render_cell("{{x}}", {"x": "1"}) == "{{x}}"

    @requires_openpyxl
    def test_non_string_cells_passthrough(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        for v in [None, 42, 3.14, True, ["a", "b"]]:
            assert engine.render_cell(v, {}) == v

    @requires_openpyxl
    def test_no_marker_means_passthrough(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        assert engine.render_cell("plain text", {}) == "plain text"
        assert engine.render_cell("", {}) == ""


class TestPhase2IfConditional:
    @requires_openpyxl
    def test_if_true_renders_value(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        result = engine.render_cell("{% if x %}yes{% endif %}", {"x": True})
        assert "yes" in result

    @requires_openpyxl
    def test_if_false_clears_cell(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        result = engine.render_cell("{% if x %}yes{% endif %}", {"x": False})
        assert "yes" not in result
        # silent → empty
        assert result.strip() == ""

    @requires_openpyxl
    def test_if_else_branch(self):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine()
        assert "A" in engine.render_cell("{% if x %}A{% else %}B{% endif %}", {"x": True})
        assert "B" in engine.render_cell("{% if x %}A{% else %}B{% endif %}", {"x": False})


class TestPhase3ForLoop:
    @requires_openpyxl
    def test_find_for_markers_simple(self, tmp_path):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        from openpyxl import Workbook
        wb = Workbook()
        s = wb.active
        s["A1"] = "{% for x in items %}"
        s["A2"] = "{{x}}"
        s["A3"] = "{% endfor %}"
        wb.save(str(tmp_path / "tpl.xlsx"))

        wb2 = load_workbook(str(tmp_path / "tpl.xlsx"))
        s2 = wb2.active
        engine = ExcelTemplateEngine()
        markers = engine.find_for_markers(s2)
        assert len(markers) == 1
        assert markers[0].var == "x"
        assert markers[0].list_expr == "items"
        assert markers[0].body_start == 2
        assert markers[0].body_end == 2
        assert markers[0].endfor_row == 3

    @requires_openpyxl
    def test_expand_single_level(self, tmp_path):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        from openpyxl import Workbook
        wb = Workbook()
        s = wb.active
        s["A1"] = "header"
        s["A2"] = "{% for x in items %}"
        s["A3"] = "{{x.number}}"
        s["B3"] = "{{x.value}}"
        s["A4"] = "{% endfor %}"
        wb.save(str(tmp_path / "tpl.xlsx"))

        wb2 = load_workbook(str(tmp_path / "tpl.xlsx"))
        s2 = wb2.active
        engine = ExcelTemplateEngine()
        items = [
            {"number": "1", "value": "first"},
            {"number": "2", "value": "second"},
            {"number": "3", "value": "third"},
        ]
        engine.expand_for_loops(s2, {"data": {}, "items": items}, {})

        assert s2["A1"].value == "header"
        assert s2["A2"].value == "1"
        assert s2["B2"].value == "first"
        assert s2["A3"].value == "2"
        assert s2["B3"].value == "second"
        assert s2["A4"].value == "3"
        assert s2["B4"].value == "third"
        # for/endfor rows removed
        assert s2["A5"].value is None
        assert s2["A6"].value is None

    @requires_openpyxl
    def test_expand_empty_list(self, tmp_path):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        from openpyxl import Workbook
        wb = Workbook()
        s = wb.active
        s["A1"] = "header"
        s["A2"] = "{% for x in items %}"
        s["A3"] = "{{x}}"
        s["A4"] = "{% endfor %}"
        s["A5"] = "footer"
        wb.save(str(tmp_path / "tpl.xlsx"))

        wb2 = load_workbook(str(tmp_path / "tpl.xlsx"))
        s2 = wb2.active
        engine = ExcelTemplateEngine()
        engine.expand_for_loops(s2, {"data": {}, "items": []}, {})

        assert s2["A1"].value == "header"
        assert s2["A2"].value == "footer"
        # for/endfor rows removed
        assert s2["A3"].value is None
        assert s2["A4"].value is None

    @requires_openpyxl
    def test_expand_uses_loop_var(self, tmp_path):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        from openpyxl import Workbook
        wb = Workbook()
        s = wb.active
        s["A1"] = "{% for x in items %}"
        s["A2"] = "{{loop.index}}/{{loop.length}} {{x}}"
        s["A3"] = "{% endfor %}"
        wb.save(str(tmp_path / "tpl.xlsx"))

        wb2 = load_workbook(str(tmp_path / "tpl.xlsx"))
        s2 = wb2.active
        engine = ExcelTemplateEngine()
        engine.expand_for_loops(s2, {"data": {}, "items": ["a", "b"]}, {})

        # 展開的 row 會 append 到 sheet 底部（原 marker rows 被刪除）
        appended_rows = [s2.cell(row=r, column=1).value for r in range(1, s2.max_row + 1)]
        assert any("1/2 a" in str(v) for v in appended_rows)
        assert any("2/2 b" in str(v) for v in appended_rows)

    @requires_openpyxl
    def test_for_marker_rows_removed_after_expansion(self, tmp_path):
        from md_word_renderer.renderer.excel_template_engine import ExcelTemplateEngine
        from openpyxl import Workbook
        wb = Workbook()
        s = wb.active
        s["A1"] = "{% for x in items %}"
        s["A2"] = "{{x}}"
        s["A3"] = "{% endfor %}"
        wb.save(str(tmp_path / "tpl.xlsx"))

        wb2 = load_workbook(str(tmp_path / "tpl.xlsx"))
        s2 = wb2.active
        engine = ExcelTemplateEngine()
        engine.expand_for_loops(s2, {"data": {}, "items": ["a"]}, {})

        # {% for %} 與 {% endfor %} 標記 cell 不應殘留
        for row in s2.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert "{%" not in cell.value


# ---------------------------------------------------- renderer integration


@pytest.fixture
def minimal_template(tmp_path):
    """最小可用的 .xlsx 樣板，含一個基本資訊 sheet 與一個 list sheet 預留"""
    wb_path = tmp_path / "tpl.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Header"
    sheet["A1"], sheet["B1"] = "field", "value"
    sheet["A2"], sheet["B2"] = "name", "{{name}}"
    sheet["A3"], sheet["B3"] = "crq", "{{crq}}"
    wb.save(wb_path)
    return wb_path


@pytest.fixture
def for_template(tmp_path):
    """含 ``{% for %}`` 標記的 list sheet 樣板"""
    wb_path = tmp_path / "tpl.xlsx"
    wb = Workbook()
    s = wb.active
    s.title = "Header"
    s["A1"], s["B1"] = "field", "value"
    s["A2"], s["B2"] = "name", "{{name}}"
    list_sheet = wb.create_sheet("items")
    list_sheet["A1"] = "number"
    list_sheet["B1"] = "value"
    list_sheet["A2"] = "{% for x in items %}"
    list_sheet["A3"] = "{{x.number}}"
    list_sheet["B3"] = "{{x.value}}"
    list_sheet["A4"] = "{% endfor %}"
    wb.save(wb_path)
    return wb_path


class TestExcelRendererCore:
    @requires_openpyxl
    def test_template_substitutes_existing_cells(self, tmp_path, minimal_template):
        """v2.2.1: 樣板內 ``{{var}}`` 應被替換"""
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer

        out = tmp_path / "out.xlsx"
        data = {"name": "Alice", "crq": "CRQ123"}
        ExcelRenderer().render_to_file(data, str(minimal_template), str(out))

        wb = load_workbook(str(out))
        sheet = wb["Header"]
        # 從 row 2 開始（row 1 是 header）
        values = {
            sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
            for r in range(2, sheet.max_row + 1)
        }
        assert values["name"] == "Alice"
        assert values["crq"] == "CRQ123"

    @requires_openpyxl
    def test_no_auto_append_for_undeclared_scalars(self, tmp_path, minimal_template):
        """v2.2.1: 資料中沒在樣板內宣告的純量欄位不應被自動 append"""
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer

        out = tmp_path / "out.xlsx"
        data = {"name": "Alice", "crq": "CRQ123", "extra1": "x", "extra2": "y"}
        ExcelRenderer().render_to_file(data, str(minimal_template), str(out))

        wb = load_workbook(str(out))
        sheet = wb["Header"]
        # 只有 row 1 (header) + row 2 (name) + row 3 (crq)
        assert sheet.max_row == 3

    @requires_openpyxl
    def test_for_loop_in_list_sheet(self, tmp_path, for_template):
        """v2.2.1: list sheet 含 ``{% for %}`` 時走 for 展開"""
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer

        out = tmp_path / "out.xlsx"
        data = {
            "name": "Alice",
            "items": [
                {"number": "1", "value": "first"},
                {"number": "2", "value": "second"},
            ],
        }
        r = ExcelRenderer()
        r.layout.template_engine.auto_flatten_lists = False
        r.render_to_file(data, str(for_template), str(out))

        wb = load_workbook(str(out))
        s = wb["items"]
        assert s["A1"].value == "number"
        assert s["A2"].value == "1"
        assert s["B2"].value == "first"
        assert s["A3"].value == "2"
        assert s["B3"].value == "second"

    @requires_openpyxl
    def test_load_template_missing_file_raises(self):
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer
        with pytest.raises(FileNotFoundError):
            ExcelRenderer().load_template("nonexistent.xlsx")

    @requires_openpyxl
    def test_factory_dispatch_by_format_hint(self, tmp_path):
        from md_word_renderer.renderer.factory import build_renderer
        tpl = tmp_path / "tpl.xlsx"
        wb = Workbook()
        wb.active.title = "Header"
        wb.save(str(tpl))

        r1 = build_renderer(template_path=str(tpl), format_hint="auto")
        r2 = build_renderer(template_path=str(tpl), format_hint="xlsx")
        assert r1.__class__.__name__ == r2.__class__.__name__ == "ExcelRenderer"

    @requires_openpyxl
    def test_template_layout_metadata_overrides(self, tmp_path):
        """LAYOUT hidden sheet 應能覆寫 layout 預設"""
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer

        wb_path = tmp_path / "with_layout.xlsx"
        wb = Workbook()
        wb.active.title = "Data"
        layout = wb.create_sheet("LAYOUT")
        layout["A1"], layout["B1"] = "key", "value"
        layout["A2"], layout["B2"] = "auto_fit_columns", False
        layout["A3"], layout["B3"] = "image_max_width_px", 200
        layout["A4"], layout["B4"] = "image_max_height_px", 150
        layout["A5"], layout["B5"] = "auto_flatten_lists", False
        wb.save(wb_path)

        out = tmp_path / "out.xlsx"
        r = ExcelRenderer()
        r.render_to_file({"A": "1"}, str(wb_path), str(out))
        assert r.layout.image.max_width_px == 200
        assert r.layout.image.max_height_px == 150
        assert r.layout.auto_fit_columns is False
        assert r.layout.template_engine.auto_flatten_lists is False
        wb2 = load_workbook(str(out))
        assert "LAYOUT" not in wb2.sheetnames

    @requires_openpyxl
    def test_existing_header_sheet_is_reused(self, tmp_path):
        """樣板內已有 header sheet 時應優先複用"""
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer

        cfg_template = tmp_path / "with_header.xlsx"
        wb = Workbook()
        wb.active.title = "Header"
        sheet = wb.active
        sheet["A1"], sheet["B1"] = "field", "value"
        sheet["A2"], sheet["B2"] = "name", "{{name}}"
        wb.save(str(cfg_template))

        out = tmp_path / "out.xlsx"
        ExcelRenderer().render_to_file({"name": "Alice"}, str(cfg_template), str(out))
        wb2 = load_workbook(str(out))
        assert "Header" in wb2.sheetnames
        sheet = wb2["Header"]
        # 樣板內的 {{name}} 應被替換
        assert sheet["B2"].value == "Alice"
        # 沒有額外建立「基本資訊」之類的 sheet
        assert len(wb2.sheetnames) == 1

    @requires_openpyxl
    def test_wrap_text_on_multiline_substituted_cell(self, tmp_path):
        from md_word_renderer.renderer.excel_renderer import ExcelRenderer

        wb_path = tmp_path / "tpl.xlsx"
        wb = Workbook()
        s = wb.active
        s.title = "Header"
        s["A1"], s["B1"] = "field", "value"
        s["A2"], s["B2"] = "notes", "{{notes}}"
        wb.save(wb_path)

        out = tmp_path / "out.xlsx"
        data = {"notes": "line1\nline2"}
        ExcelRenderer().render_to_file(data, str(wb_path), str(out))
        wb2 = load_workbook(str(out))
        sheet = wb2["Header"]
        cell = sheet["B2"]
        assert cell.value == "line1\nline2"
        assert cell.alignment.wrap_text is True
