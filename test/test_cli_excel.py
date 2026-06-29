# -*- coding: utf-8 -*-
"""
測試 CLI 在 Excel 樣板下的行為
"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))


try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


requires_openpyxl = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl 不可用")


# --------------------------------------------------------------- fixtures


@pytest.fixture
def md_file(tmp_path):
    p = tmp_path / "data.md"
    p.write_text(
        "# sample\n"
        "\n"
        "1. system_name | example\n"
        "2. crq | CRQ001\n"
        "3. cases |\n"
        "    1. item A\n"
        "       1. sub 1\n"
        "    2. item B\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def xlsx_template(tmp_path):
    """v2.2.1 樣板：含 ``{{var}}`` 與 ``{% for %}`` 標記"""
    p = tmp_path / "tpl.xlsx"
    wb = Workbook()
    s = wb.active
    s.title = "Header"
    s["A1"], s["B1"] = "field", "value"
    s["A2"], s["B2"] = "system_name", "{{system_name}}"
    s["A3"], s["B3"] = "crq", "{{crq}}"
    list_sheet = wb.create_sheet("items")
    list_sheet["A1"] = "number"
    list_sheet["B1"] = "value"
    list_sheet["A2"] = "{% for x in items %}"
    list_sheet["A3"] = "{{x.number}}"
    list_sheet["B3"] = "{{x.value}}"
    list_sheet["A4"] = "{% endfor %}"
    wb.save(p)
    return p


# --------------------------------------------------------------- CLI


class TestRenderCommandXlsx:
    @requires_openpyxl
    def test_render_xlsx(self, tmp_path, md_file, xlsx_template):
        from md_word_renderer.cli import cli

        out = tmp_path / "out.xlsx"
        exit_code = cli([
            "render",
            str(md_file),
            str(xlsx_template),
            str(out),
            "--no-validate",
        ])
        assert exit_code == 0
        assert out.exists()

        # Verify content（v2.2.1：樣板為主，{{}} 已替換）
        wb = load_workbook(str(out))
        assert "Header" in wb.sheetnames
        sheet = wb["Header"]
        # header + 2 替換列
        assert sheet.max_row >= 3
        values = {
            sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
            for r in range(2, sheet.max_row + 1)
        }
        assert values.get("system_name") == "example"
        assert values.get("crq") == "CRQ001"

    @requires_openpyxl
    def test_render_with_explicit_format_xlsx(self, tmp_path, md_file, xlsx_template):
        from md_word_renderer.cli import cli

        out = tmp_path / "out.xlsx"
        exit_code = cli([
            "render",
            str(md_file),
            str(xlsx_template),
            str(out),
            "--format", "xlsx",
            "--no-validate",
        ])
        assert exit_code == 0
        assert out.exists()

    @requires_openpyxl
    def test_render_format_help_lists_xlsx_choice(self):
        from md_word_renderer.cli import create_parser
        p = create_parser()
        help_text = p.format_help()
        assert "--format" in help_text
        assert "xlsx" in help_text
        assert "docx" in help_text


class TestBatchCommandXlsx:
    @requires_openpyxl
    def test_batch_with_xlsx(self, tmp_path, md_file, xlsx_template):
        from md_word_renderer.cli import cli

        in_dir = tmp_path / "in"
        in_dir.mkdir()
        for stem in ("a", "b"):
            (in_dir / f"{stem}.md").write_text(md_file.read_text(encoding="utf-8"), encoding="utf-8")

        out_dir = tmp_path / "out"
        exit_code = cli([
            "batch",
            str(in_dir),
            str(xlsx_template),
            str(out_dir),
            "--no-validate",
        ])
        assert exit_code == 0
        produced = list(out_dir.glob("*.xlsx"))
        assert len(produced) == 2


class TestBatchTemplatesXlsx:
    @requires_openpyxl
    def test_batch_templates_xlsx_outputs_preserve_extension(self, tmp_path, md_file):
        from md_word_renderer.cli import cli

        tpl_dir = tmp_path / "templates"
        tpl_dir.mkdir()
        for name in ("t1", "t2"):
            wb = Workbook()
            wb.active.title = "Header"
            wb.active["A1"], wb.active["B1"] = "field", "value"
            wb.save(tpl_dir / f"{name}.xlsx")

        out_dir = tmp_path / "out"
        exit_code = cli([
            "batch-templates",
            str(md_file),
            str(tpl_dir),
            str(out_dir),
            "--no-validate",
        ])
        assert exit_code == 0
        produced = sorted(p.name for p in out_dir.iterdir())
        assert produced == ["t1.xlsx", "t2.xlsx"]


class TestExitCodes:
    @requires_openpyxl
    def test_render_xlsx_missing_template_returns_nonzero(self, tmp_path, md_file):
        from md_word_renderer.cli import cli

        out = tmp_path / "out.xlsx"
        exit_code = cli([
            "render",
            str(md_file),
            str(tmp_path / "no_such_template.xlsx"),
            str(out),
            "--no-validate",
        ])
        assert exit_code != 0
