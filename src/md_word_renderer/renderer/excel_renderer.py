"""
Excel 渲染器（v2.2.1）

設計（template-driven，v2.2.1 起）：
1. 樣板為主：cell 內的 ``{{...}}`` / ``{% if %}`` / ``{% for %}`` 走 Jinja2 引擎
2. 純量欄位不再自動 append；只在樣板有 ``{{...}}`` 時替換
3. 頂層 list 欄位：若樣板有對應 sheet 且含 ``{% for %}`` → 展開；否則依 ``auto_flatten_lists`` 決定是否 auto-flatten
4. 樣板若提供隱藏 ``LAYOUT`` 工作表可覆寫 ``LayoutConfig`` 設定
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]
    Worksheet = None  # type: ignore[assignment]
    HAS_OPENPYXL = False

from .error_handler import RenderErrorHandler
from .excel_layout import LayoutConfig, sanitize_sheet_name
from .excel_image_handler import ExcelImageHandler, ExcelImageError
from .excel_template_engine import ExcelTemplateEngine


_EXCEL_MAX_COL_WIDTH = 100.0


class ExcelRenderError(Exception):
    """Excel 渲染錯誤"""


def _is_scalar(v: Any) -> bool:
    return isinstance(v, (str, int, float, bool)) or v is None


def _coerce_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


class ExcelRenderer:
    """
    Excel 渲染器（與 ``WordRenderer`` 同形 API）
    """

    DEFAULT_HEADER_ROW = ["欄位", "值"]
    DEFAULT_LIST_HEADER_ROW = ["項目編號", "值", "類型"]
    LIST_SHEET_METADATA_NAME = "LAYOUT"

    def __init__(
        self,
        layout: Optional[LayoutConfig] = None,
        show_errors: bool = True,
        error_format: str = "[ERROR: 變數 '{var}' 不存在]",
    ):
        if not HAS_OPENPYXL:
            raise ExcelRenderError(
                "openpyxl 未安裝；請執行 `pip install openpyxl` 後再使用"
            )
        self.layout = layout or LayoutConfig()
        self.template_path: Optional[str] = None
        self.workbook = None
        self.error_handler = RenderErrorHandler(
            show_errors=show_errors, error_format=error_format
        )
        self.show_errors = show_errors
        self.error_format = error_format
        self.image_handler = ExcelImageHandler(
            max_width_px=self.layout.image.max_width_px,
            max_height_px=self.layout.image.max_height_px,
        )
        self.engine = ExcelTemplateEngine(
            enabled=self.layout.template_engine.enabled,
            missing_variable=self.layout.template_engine.missing_variable,
            error_format=self.layout.template_engine.error_format,
        )

    # ------------------------------------------------------------- public API

    def load_template(self, template_path: str) -> None:
        path = Path(template_path)
        if not path.exists():
            raise FileNotFoundError(f"模板檔案不存在: {template_path}")

        self.template_path = str(path)
        self.workbook = load_workbook(str(path))
        self._apply_template_metadata()

    def render(self, data: Dict[str, Any]) -> None:
        if self.workbook is None:
            raise ExcelRenderError("請先使用 load_template() 載入樣板")

        processed = self._prepare_context(data)
        context = self._build_template_context(processed)
        existing_sheets = set(self.workbook.sheetnames)
        used_sheet_names: set = set()

        # 1. 標頭 sheet：只套模板，不 append
        if self.layout.header_sheet.enabled:
            header_name = self._resolve_header_sheet_name(used_sheet_names, existing_sheets)
            self._ensure_header_sheet(header_name)
            used_sheet_names.add(header_name)

        # 2. 頂層 list 欄位的處理（含兩種路徑）
        for key, value in processed.items():
            if key == "data":
                continue
            if not isinstance(value, list):
                continue
            self._render_or_expand_list_sheet(
                key, value, context, processed, used_sheet_names, existing_sheets
            )

        # 2.5 對所有「未對應到資料 key」的既有 sheet 也走模板（支援第二張以上的自訂 sheet）
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in used_sheet_names:
                continue
            if sheet_name == self.LIST_SHEET_METADATA_NAME:
                continue
            sheet = self.workbook[sheet_name]
            if self.engine.has_for_marker(sheet):
                self.engine.expand_for_loops(sheet, context, processed)
                used_sheet_names.add(sheet_name)
            elif self.layout.template_engine.auto_flatten_lists:
                # 對應不到資料 key、沒 for marker、auto_flatten=true → 不建立新 sheet（既有的沒資料）
                # 略過
                pass

        # 3. 對所有 sheet 套模板 pass（含 cell 內的 {{var}}）
        self._apply_template_pass_to_all(processed)

        if self.layout.auto_fit_columns:
            for sheet_name in self.workbook.sheetnames:
                self._auto_fit_columns(self.workbook[sheet_name])

    def save(self, output_path: str) -> None:
        if self.workbook is None:
            raise ExcelRenderError("請先載入並渲染樣板")

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        try:
            self.workbook.save(str(output))
        except Exception as exc:
            raise ExcelRenderError(f"儲存失敗: {exc}") from exc

    def render_to_file(
        self,
        data: Dict[str, Any],
        template_path: str,
        output_path: str,
    ) -> None:
        self.load_template(template_path)
        self.render(data)
        self.save(output_path)

    # --------------------------------------------------------------- context

    def _prepare_context(self, data: Dict[str, Any]) -> Dict[str, Any]:
        return data

    def _build_template_context(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """對齊 WordRenderer._prepare_context：頂層欄位直接取、`data` 鍵可取特殊字元"""
        context = dict(data)
        context["data"] = data
        return context

    def _apply_template_pass_to_all(self, data: Dict[str, Any]) -> None:
        """對 workbook 中所有 sheet 跑 Jinja2 模板 pass

        略過 LAYOUT metadata sheet（已在 _apply_template_metadata 移除，
        此為保險）。注入 _image_handler 給 context，讓 for 內 image 嵌入能運作。
        """
        if not self.engine.enabled or not self.layout.template_engine.enabled:
            return
        context = self._build_template_context(data)
        context["_image_handler"] = self.image_handler
        for sheet_name in self.workbook.sheetnames:
            if sheet_name == self.LIST_SHEET_METADATA_NAME:
                continue
            self.engine.render_sheet(self.workbook[sheet_name], context)

    # ----------------------------------------------------------- sheet helpers

    def _apply_template_metadata(self) -> None:
        if not self.workbook:
            return
        if self.LIST_SHEET_METADATA_NAME not in self.workbook.sheetnames:
            return

        meta_sheet = self.workbook[self.LIST_SHEET_METADATA_NAME]
        meta_dict: Dict[str, Any] = {}

        for row in meta_sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            value = row[1] if len(row) > 1 else None
            meta_dict[key] = value

        merged: Dict[str, Any] = {
            "default_template": meta_dict.get("default_template", self.layout.default_template),
            "list_sheet_naming": meta_dict.get("list_sheet_naming", self.layout.list_sheet_naming),
            "header_sheet": {
                "enabled": self._as_bool(meta_dict.get("header_enabled"), default=self.layout.header_sheet.enabled),
                "name": meta_dict.get("header_name", self.layout.header_sheet.name) or self.layout.header_sheet.name,
            },
            "extra_columns": meta_dict.get("extra_columns", self.layout.extra_columns),
            "image": {
                "max_width_px": self._as_int(meta_dict.get("image_max_width_px"), default=self.layout.image.max_width_px),
                "max_height_px": self._as_int(meta_dict.get("image_max_height_px"), default=self.layout.image.max_height_px),
            },
            "auto_fit_columns": self._as_bool(meta_dict.get("auto_fit_columns"), default=self.layout.auto_fit_columns),
            "template_engine": {
                "enabled": self._as_bool(meta_dict.get("template_engine_enabled"), default=self.layout.template_engine.enabled),
                "auto_flatten_lists": self._as_bool(meta_dict.get("auto_flatten_lists"), default=self.layout.template_engine.auto_flatten_lists),
                "missing_variable": str(meta_dict.get("missing_variable", self.layout.template_engine.missing_variable)),
                "error_format": str(meta_dict.get("error_format", self.layout.template_engine.error_format)),
            },
        }

        self.layout = LayoutConfig.from_mapping(merged)
        self.image_handler = ExcelImageHandler(
            max_width_px=self.layout.image.max_width_px,
            max_height_px=self.layout.image.max_height_px,
        )
        self.engine = ExcelTemplateEngine(
            enabled=self.layout.template_engine.enabled,
            missing_variable=self.layout.template_engine.missing_variable,
            error_format=self.layout.template_engine.error_format,
        )

        try:
            del self.workbook[self.LIST_SHEET_METADATA_NAME]
        except Exception:
            pass

    @staticmethod
    def _as_bool(value: Any, default: bool = False) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"true", "1", "yes", "y", "on"}:
            return True
        if text in {"false", "0", "no", "n", "off", ""}:
            return False
        return default

    @staticmethod
    def _as_int(value: Any, default: int) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def _resolve_header_sheet_name(self, used: set, existing_sheets: set = None) -> str:
        base = sanitize_sheet_name(self.layout.header_sheet.name, fallback="Header")
        existing_sheets = existing_sheets or set()

        for existing in existing_sheets:
            if existing == base and existing not in used:
                return existing

        if len(existing_sheets) == 1:
            only = next(iter(existing_sheets))
            if only not in used:
                return only

        if base not in used:
            return base
        counter = 2
        candidate = base
        while candidate in used:
            suffix = f"_{counter}"
            candidate = f"{base[: 31 - len(suffix)]}{suffix}"
            counter += 1
        return candidate

    def _ensure_header_sheet(self, sheet_name: str) -> None:
        """確保 header sheet 存在；不再 append 純量資料。

        若樣板已有該 sheet → 沿用；否則建立空 sheet 並補上預設標題列。
        """
        if sheet_name in self.workbook.sheetnames:
            return
        sheet = self.workbook.create_sheet(sheet_name)
        self._ensure_header_row(sheet, self.DEFAULT_HEADER_ROW)

    def _render_or_expand_list_sheet(
        self,
        key: str,
        items: List[Dict[str, Any]],
        context: Dict[str, Any],
        data: Dict[str, Any],
        used_sheet_names: set,
        existing_sheets: set,
    ) -> None:
        if key in existing_sheets:
            sheet = self.workbook[key]
            if self.engine.has_for_marker(sheet):
                # 走 for 展開（auto_flatten 在此 sheet 不啟用）
                self.engine.expand_for_loops(sheet, context, data)
                return
            if self.layout.template_engine.auto_flatten_lists:
                # 向後相容：v2.2 行為（auto-flatten）
                self._render_list_sheet_flatten(key, items, sheet)
                return
            # auto_flatten_lists=false → 樣板沒寫 for、沒啟用 auto_flatten → 略過
            return

        # 樣板無對應 sheet
        if self.layout.template_engine.auto_flatten_lists:
            sheet_name = self.layout.resolve_list_sheet_name(key, used_sheet_names)
            self._render_list_sheet_create(sheet_name, key, items)
            used_sheet_names.add(sheet_name)

    def _render_list_sheet_flatten(
        self, key: str, items: List[Dict[str, Any]], sheet: Worksheet
    ) -> None:
        headers = list(self.DEFAULT_LIST_HEADER_ROW) + list(self.layout.extra_columns)
        self._ensure_header_row(sheet, headers)
        for item in items:
            self._flatten_into_sheet(sheet, key, item, path_so_far="")

    def _render_list_sheet_create(
        self, sheet_name: str, key: str, items: List[Dict[str, Any]]
    ) -> None:
        sheet = self.workbook.create_sheet(sheet_name)
        self._render_list_sheet_flatten(key, items, sheet)

    def _flatten_into_sheet(
        self,
        sheet: Worksheet,
        field_key: str,
        item: Dict[str, Any],
        path_so_far: str,
    ) -> None:
        children = item.get("children") or []
        number = item.get("number") or ""
        value = item.get("value", "")
        item_type = item.get("type", "text")

        current_path = f"{path_so_far}{number}" if number else path_so_far
        depth = max(0, current_path.count(".")) if current_path else 0

        if not children:
            row = sheet.max_row + 1 if sheet.max_row else 2
            sheet.cell(row=row, column=1, value=number)
            value_cell = sheet.cell(row=row, column=2, value=_coerce_str(value))
            sheet.cell(row=row, column=3, value=item_type)
            if "\n" in (value or ""):
                value_cell.alignment = self._wrap_alignment()

            extra_start = 4
            for offset, col_name in enumerate(self.layout.extra_columns):
                sheet.cell(row=row, column=extra_start + offset, value=self._extra_column_value(col_name, field_key, current_path, depth, item))

            if item_type == "image":
                image_path = item.get("image_path")
                if image_path:
                    try:
                        self.image_handler.embed(
                            worksheet=sheet,
                            cell_ref=f"B{row}",
                            image_path=image_path,
                            alt_text=item.get("image_alt"),
                        )
                    except ExcelImageError as exc:
                        sheet.cell(row=row, column=3, value=f"image-missing: {exc}")
            return

        row = sheet.max_row + 1 if sheet.max_row else 2
        if number:
            sheet.cell(row=row, column=1, value=number)
            value_cell = sheet.cell(row=row, column=2, value=_coerce_str(value))
            sheet.cell(row=row, column=3, value="group" if item_type == "text" else item_type)
            if "\n" in (value or ""):
                value_cell.alignment = self._wrap_alignment()

            extra_start = 4
            for offset, col_name in enumerate(self.layout.extra_columns):
                sheet.cell(row=row, column=extra_start + offset, value=self._extra_column_value(col_name, field_key, current_path, depth, item))

        for child in children:
            self._flatten_into_sheet(sheet, field_key, child, path_so_far=current_path + "." if current_path else "")

    @staticmethod
    def _extra_column_value(
        col_name: str,
        field_key: str,
        path: str,
        depth: int,
        item: Dict[str, Any],
    ) -> Any:
        if col_name == "source_field":
            return field_key
        if col_name == "path":
            return path or ""
        if col_name == "depth":
            return depth
        if col_name == "type":
            return item.get("type", "text")
        return item.get(col_name, "")

    @staticmethod
    def _wrap_alignment():
        from openpyxl.styles import Alignment
        return Alignment(wrap_text=True)

    def _ensure_header_row(self, sheet: Worksheet, headers: List[str]) -> None:
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        first_non_empty = False
        if first_row:
            for value in first_row[0]:
                if value not in (None, ""):
                    first_non_empty = True
                    break
        if first_non_empty:
            return
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

    @staticmethod
    def _auto_fit_columns(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            if not column_cells:
                continue
            column_letter = column_cells[0].column_letter  # type: ignore[attr-defined]
            max_len = 0
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                length = max((len(str(line)) for line in str(value).split("\n")), default=0)
                if length > max_len:
                    max_len = length
            adjusted = min(max(8.0, float(max_len) + 2.0), _EXCEL_MAX_COL_WIDTH)
            sheet.column_dimensions[column_letter].width = adjusted


# ----------------------------------------------------------------- helpers


def flatten_markdown_data(data: Dict[str, Any]) -> Dict[str, Any]:
    scalars: Dict[str, Any] = {}
    lists: Dict[str, List[Any]] = {}
    for key, value in data.items():
        if key == "data":
            continue
        if isinstance(value, list):
            lists[key] = value
        else:
            scalars[key] = value
    return {"scalars": scalars, "lists": lists}
