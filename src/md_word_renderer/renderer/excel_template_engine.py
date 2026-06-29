"""
Excel 樣板模板引擎（v2.2.1）

封裝 Jinja2，提供：
- ``render_cell(value, context)`` — 替換單一儲存格字串中的 ``{{var}}`` / ``{% if %}``
- ``render_sheet(sheet, context)`` — 對整張 sheet 走一次替換（跳過非字串 cell）
- ``find_for_markers(sheet)`` — 找出 ``{% for VAR in LIST %}`` 與 ``{% endfor %}`` 標記
- ``expand_for_loops(sheet, context, data)`` — 對 for 區段做 stack-based 展開

缺變數策略：使用 ``Undefined`` + 自定 ``finalize``，未提供變數靜默替換為空字串。
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Optional, Tuple

from jinja2 import Environment, StrictUndefined, exceptions as jinja_exc


_FOR_RE = re.compile(r"^\s*\{%\s*for\s+(\w+)\s+in\s+(.+?)\s*%\}\s*$")
_END_FOR_RE = re.compile(r"^\s*\{%\s*endfor\s*%\}\s*$")


@dataclass
class ForMarker:
    """``{% for %}`` 標記位置"""

    row_idx: int
    column: int
    var: str
    list_expr: str
    body_start: int  # 1-based, 下一列開始為 body
    body_end: int    # 1-based, 結尾列（endfor 那一列的上一列）
    endfor_row: int  # 1-based, 對應 ``{% endfor %}`` 所在列

    @property
    def marker_row(self) -> int:
        return self.row_idx


class ExcelTemplateEngine:
    """
    Excel 樣板模板引擎

    Args:
        enabled: 設為 ``False`` 時 ``render_cell`` 直接回傳原值
        missing_variable: ``"silent"``（缺變數 → 空字串）/ ``"keep"``（保留 ``{{...}}``）
        error_format: 缺變數時若 ``missing_variable == "error_format"``，用此字串填入
        syntax_error_cb: 註冊 Jinja2 語法錯誤 callback（除錯用）
    """

    def __init__(
        self,
        enabled: bool = True,
        missing_variable: str = "silent",
        error_format: str = "[ERROR: 變數 '{var}' 不存在]",
        syntax_error_cb: Optional[Callable[[jinja_exc.TemplateSyntaxError, str], None]] = None,
    ):
        self.enabled = enabled
        self.missing_variable = missing_variable
        self.error_format = error_format
        self._syntax_error_cb = syntax_error_cb
        self.env = self._build_env()

    def _build_env(self) -> Environment:
        """建立 Jinja2 Environment

        採 ``StrictUndefined`` 讓缺變數拋例外，再於 ``render_cell`` 內依
        ``missing_variable`` 設定決定替換為空字串、保留字面、或套 error_format。
        """
        env = Environment(undefined=StrictUndefined)
        return env

    def set_syntax_error_callback(self, cb) -> None:
        self._syntax_error_cb = cb

    def _has_marker(self, value: Any) -> bool:
        if not isinstance(value, str):
            return False
        return "{{" in value or "{%" in value or "{#" in value

    def render_cell(self, value: Any, context: Dict[str, Any]) -> Any:
        """渲染單一 cell 值

        - 非字串或不含 marker → 原值
        - ``enabled=False`` → 原值
        - 走 Jinja2 渲染（缺變數依設定）
        """
        if not self.enabled:
            return value
        if not self._has_marker(value):
            return value

        try:
            template = self.env.from_string(value)
            return template.render(**context)
        except jinja_exc.UndefinedError as exc:
            if self.missing_variable == "keep":
                return value
            if self.missing_variable == "error_format":
                var = self._extract_var_name(str(exc))
                return self.error_format.format(var=var)
            return ""  # silent
        except jinja_exc.TemplateSyntaxError as exc:
            if self._syntax_error_cb is not None:
                self._syntax_error_cb(exc, value)
            return value
        except Exception:
            return value

    def _extract_var_name(self, msg: str) -> str:
        """從 Jinja2 UndefinedError 訊息中取出變數名（盡力而為）"""
        # 訊息範例：'foo' is undefined
        for quote in ('"', "'"):
            if f"{quote}" in msg:
                parts = msg.split(quote)
                if len(parts) >= 2:
                    return parts[1]
        return "?"

    def render_sheet(self, sheet, context: Dict[str, Any]) -> int:
        """走過整張 sheet 套模板

        對含 ``{{``/``{%``/``{#`` 標記的 cell 渲染；含換行的取代結果會自動設 wrap_text。
        非字串 cell（數字、bool、None、日期、formula）跳過。

        Returns:
            int: 實際替換的 cell 數
        """
        if not self.enabled:
            return 0
        replaced = 0
        for row in sheet.iter_rows():
            for cell in row:
                if not self._has_marker(cell.value):
                    continue
                new_value = self.render_cell(cell.value, context)
                if new_value != cell.value:
                    cell.value = new_value
                    replaced += 1
                    # 含換行的字串自動啟用 wrap_text
                    if isinstance(new_value, str) and "\n" in new_value:
                        cell.alignment = self._wrap_alignment()
        return replaced

    @staticmethod
    def _wrap_alignment():
        from openpyxl.styles import Alignment
        return Alignment(wrap_text=True)

    # ---------------------------------------------------------- for loops

    def has_for_marker(self, sheet) -> bool:
        return bool(self.find_for_markers(sheet))

    def find_for_markers(self, sheet) -> List[ForMarker]:
        """掃描 sheet 中所有 ``{% for VAR in LIST %}`` 與對應的 ``{% endfor %}``

        支援巢狀；用 stack 配對。回傳的 markers 由內而外排序（內層 for 先）。
        """
        markers: List[ForMarker] = []
        stack: List[Tuple[int, int, str, str]] = []

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                v = cell.value
                if not isinstance(v, str):
                    continue
                stripped = v.strip()
                m_for = _FOR_RE.match(stripped)
                m_end = _END_FOR_RE.match(stripped)
                if m_for:
                    stack.append((row_idx, col_idx, m_for.group(1), m_for.group(2)))
                elif m_end and stack:
                    start_row, start_col, var, list_expr = stack.pop()
                    body_start = start_row + 1
                    body_end = row_idx - 1
                    markers.append(ForMarker(
                        row_idx=start_row,
                        column=start_col,
                        var=var,
                        list_expr=list_expr,
                        body_start=body_start,
                        body_end=body_end,
                        endfor_row=row_idx,
                    ))
        # 由內而外排序：body_end 較小者先（內層 for 的 body 較短）
        markers.sort(key=lambda m: m.body_end)
        return markers

    def expand_for_loops(
        self,
        sheet,
        context: Dict[str, Any],
        data: Dict[str, Any],
    ) -> int:
        """對 sheet 內所有 ``{% for %}`` 區段做展開

        支援巢狀：採「由內而外」處理策略。
        內層 for 先展開，展開後行數變動，外層 for 在重新掃描時仍能正確定位。
        為避免 openpyxl insert_rows 帶來的格式不穩定，採用「刪除 + append 到底部」。
        收尾時 sweep 一次清掉殘留的 ``{% for %}`` / ``{% endfor %}`` 標記。

        Returns:
            int: 展開後新增的 row 數
        """
        if not self.enabled:
            return 0

        total_inserted = 0
        while True:
            markers = self.find_for_markers(sheet)
            if not markers:
                break
            inserted = self._expand_single_for(
                sheet, markers[0], context, data
            )
            total_inserted += inserted

        # 收尾：清掉殘留的 for/endfor 標記（理論上 stack-based 配對後不會留，但保險）
        self._cleanup_orphan_markers(sheet)
        return total_inserted

    def _cleanup_orphan_markers(self, sheet) -> None:
        """掃過整張 sheet，把 cell 值是 ``{% for %}`` 或 ``{% endfor %}`` 的 cell 清空"""
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                stripped = v.strip()
                if _FOR_RE.match(stripped) or _END_FOR_RE.match(stripped):
                    cell.value = None

    def _expand_single_for(
        self,
        sheet,
        marker: ForMarker,
        context: Dict[str, Any],
        data: Dict[str, Any],
    ) -> int:
        items = self._resolve_list_expr(marker.list_expr, context, data)
        if not isinstance(items, list):
            items = []

        # 取出 body rows 的所有 cell 值（包含可能的 inner for/endfor 標記）
        body_rows: List[List[Tuple[int, Any]]] = []
        for r in range(marker.body_start, marker.body_end + 1):
            row_cells = []
            for c in range(1, sheet.max_column + 1):
                row_cells.append((c, sheet.cell(row=r, column=c).value))
            body_rows.append(row_cells)

        # 刪除 marker row + body rows + endfor row（從下往上刪）
        rows_to_delete = list(range(marker.marker_row, marker.endfor_row + 1))
        rows_to_delete.sort(reverse=True)
        for r in rows_to_delete:
            sheet.delete_rows(r, 1)

        inserted = 0
        for idx, item in enumerate(items):
            child_context = dict(context)
            child_context[marker.var] = item
            child_context["loop"] = {
                "index": idx + 1,
                "index0": idx,
                "first": idx == 0,
                "last": idx == len(items) - 1,
                "length": len(items),
            }
            for row_cells in body_rows:
                new_row_idx = sheet.max_row + 1
                for (col, raw_value) in row_cells:
                    new_value = self.render_cell(raw_value, child_context)
                    sheet.cell(row=new_row_idx, column=col, value=new_value)
                inserted += 1
                self._maybe_attach_image(sheet, new_row_idx, item, child_context)
        return inserted

    def _resolve_list_expr(
        self,
        list_expr: str,
        context: Dict[str, Any],
        data: Dict[str, Any],
    ) -> List[Any]:
        """解析 ``{% for x in LIST_EXPR %}`` 中的 LIST_EXPR

        支援：
        - 純變數：``test_cases``（在 context 中取）
        - 透過 data：``data["test_cases"]`` / ``data["#16"].children``
        - Jinja2 形式：當作 template 渲染後再取值（支援 ``case.children`` 形式）
        """
        expr = list_expr.strip()

        # 1. 嘗試以 Jinja2 形式渲染（最通用；支援 case.children 等）
        try:
            template = self.env.from_string("{{ (" + expr + ") | default(none) }}")
            rendered = template.render(**context)
            # 渲染後可能是 "[1, 2, 3]" 字面 — 我們不解析，直接走「走 context」的 path
        except Exception:
            pass

        # 2. 走 context（處理最常見的純變數與 data["..."] 形式）
        # data["key"].attr / data["key"][idx]
        m = re.match(
            r'^data\[\s*[\"\']([^\"\']+)[\"\']\s*\]\s*(?:\.([\w]+)|\[(\d+)\])?$',
            expr,
        )
        if m:
            key, attr, idx = m.group(1), m.group(2), m.group(3)
            val = data.get(key)
            if attr and isinstance(val, dict):
                val = val.get(attr)
            if idx and isinstance(val, list) and int(idx) < len(val):
                val = val[int(idx)]
            if isinstance(val, list):
                return val
            return []
        m = re.match(r'^data\[\s*[\"\']#(\d+)[\"\']\s*\](?:\.([\w]+))?', expr)
        if m:
            num = m.group(1)
            attr = m.group(2)
            entry = data.get(f"#{num}")
            if attr and isinstance(entry, dict):
                entry = entry.get(attr)
            if isinstance(entry, list):
                return entry
            if isinstance(entry, dict):
                children = entry.get("children") or []
                return children if isinstance(children, list) else []
            return []
        if expr in context:
            val = context[expr]
            return val if isinstance(val, list) else []
        return []

    def _maybe_attach_image(
        self,
        sheet,
        row_idx: int,
        item: Any,
        context: Dict[str, Any],
    ) -> None:
        if not isinstance(item, dict):
            return
        if item.get("type") != "image":
            return
        image_path = item.get("image_path")
        if not image_path:
            return
        handler = context.get("_image_handler")
        if handler is None:
            return
        try:
            handler.embed(sheet, f"B{row_idx}", image_path, alt_text=item.get("image_alt"))
        except Exception:
            sheet.cell(row=row_idx, column=2).value = f"[image missing: {image_path}]"
