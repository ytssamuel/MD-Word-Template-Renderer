#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
建立 Excel 樣板範例（v2.2.1 — 樣板為主，Jinja2 驅動）

產出：
- ``templates/excel/sample_template.xlsx`` — 基本 ``{{var}}`` 替換示範
- ``templates/excel/with_lists_template.xlsx`` — ``{% for %}`` 列展開示範
- ``templates/excel/with_layout_template.xlsx`` — 含 LAYOUT metadata 覆寫

執行：``python scripts/create_excel_templates.py``
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (ValueError, AttributeError):
        pass


EXCEL_DIR = Path("templates/excel")


HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="微軟正黑體", color="FFFFFFFF", size=12, bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_row(sheet, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _set_widths(sheet, widths):
    for idx, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = w


def create_sample_template():
    """最簡樣板：header + 幾個 ``{{var}}`` 替換示範（樣板為主，不會自動 append 純量欄位）"""
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "基本資訊"
    _style_header_row(sheet, ["欄位", "值"])
    _set_widths(sheet, [24, 60])

    # 純量替換
    sheet["A2"], sheet["B2"] = "系統名稱", "{{系統名稱}}"
    sheet["A3"], sheet["B3"] = "變更單號", "{{變更單號}}"
    sheet["A4"], sheet["B4"] = "預定換版日期", "{{預定換版日期}}"
    sheet["A5"], sheet["B5"] = "預定變更時段", "{{預定變更時段}}"
    sheet["A6"], sheet["B6"] = "備援路徑", "{{備援路徑}}"
    # 特殊字元鍵（括號、斜線）走 data["..."] 形式
    sheet["A7"], sheet["B7"] = "需求依據(INC/PBI)", '{{data["需求依據(INC/PBI)"]}}'
    sheet["A8"], sheet["B8"] = "通知/公告方式", '{{data["通知/公告方式"]}}'
    # 條件渲染
    sheet["A9"], sheet["B9"] = "資料庫", '{% if 資料庫 %}有{{ 資料庫 }}{% else %}無{% endif %}'
    # 編號索引
    sheet["A10"], sheet["B10"] = "第一個欄位的 key", "{{data['#1'].key}}"

    output = EXCEL_DIR / "sample_template.xlsx"
    wb.save(output)
    return output


def create_with_lists_template():
    """含 ``{% for %}`` 列展開示範（單層）"""
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- 基本資訊 sheet（替換 + if）---
    sheet = wb.active
    sheet.title = "基本資訊"
    _style_header_row(sheet, ["欄位", "值"])
    _set_widths(sheet, [24, 60])
    sheet["A2"], sheet["B2"] = "系統名稱", "{{系統名稱}}"
    sheet["A3"], sheet["B3"] = "變更單號", "{{變更單號}}"

    # --- 異動內容-測試案例 sheet（for 迴圈示範）---
    list_sheet = wb.create_sheet("異動內容-測試案例")
    _style_header_row(list_sheet, ["項目編號", "內容", "型別", "子項數"])
    _set_widths(list_sheet, [10, 60, 12, 12])

    # row 2: for marker
    list_sheet["A2"] = "{% for case in data['#16'].children %}"
    # row 3: case 本身的內容（單層，children 數量以 length 表示）
    list_sheet["A3"] = "{{case.number}}"
    list_sheet["B3"] = "{{case.value}}"
    list_sheet["C3"] = "{{case.type or 'text'}}"
    list_sheet["D3"] = "{{case.children|length}}"
    # row 4: endfor
    list_sheet["A4"] = "{% endfor %}"

    # --- 異動內容子項 sheet（攤平 children 為另一張表）---
    # v2.2.1 支援單層 ``{% for %}``；巢狀 for 為 v2.2.2+ 規劃。
    # 此 sheet 示範「用樣板逐 case 展開，每個 case 透過 child position 取子項」
    sub_sheet = wb.create_sheet("異動內容-子項")
    _style_header_row(sub_sheet, ["父項編號", "父項內容", "子項數", "第一個子項編號", "第一個子項值"])
    _set_widths(sub_sheet, [10, 50, 10, 14, 60])
    # row 2: outer for marker
    sub_sheet["A2"] = "{% for case in data['#16'].children %}"
    # row 3: 父項資料
    sub_sheet["A3"] = "{{case.number}}"
    sub_sheet["B3"] = "{{case.value}}"
    sub_sheet["C3"] = "{{case.children|length}}"
    sub_sheet["D3"] = "{% if case.children %}{{case.children[0].number}}{% endif %}"
    sub_sheet["E3"] = "{% if case.children %}{{case.children[0].value}}{% endif %}"
    # row 4: outer endfor
    sub_sheet["A4"] = "{% endfor %}"

    # --- LAYOUT metadata sheet ---
    layout = wb.create_sheet("LAYOUT")
    layout["A1"], layout["B1"] = "key", "value"
    rows = [
        ("default_template", "templates/excel/with_lists_template.xlsx"),
        ("list_sheet_naming", "{key}"),
        ("header_enabled", True),
        ("header_name", "基本資訊"),
        ("auto_fit_columns", True),
        ("image_max_width_px", 480),
        ("image_max_height_px", 360),
        ("extra_columns", "source_field,path,depth"),
        ("template_engine_enabled", True),
        ("auto_flatten_lists", False),
        ("missing_variable", "silent"),
        ("error_format", "[ERROR: 變數 '{var}' 不存在]"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        layout.cell(row=i, column=1, value=k)
        layout.cell(row=i, column=2, value=v)
    layout.column_dimensions["A"].width = 28
    layout.column_dimensions["B"].width = 40
    layout.sheet_state = "hidden"

    output = EXCEL_DIR / "with_lists_template.xlsx"
    wb.save(output)
    return output


def create_with_layout_template():
    """含 LAYOUT metadata 的最簡樣板（示範從 metadata 讀設定）"""
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Header"
    _style_header_row(sheet, ["field", "value"])
    _set_widths(sheet, [24, 60])
    sheet["A2"], sheet["B2"] = "system_name", "{{系統名稱}}"
    sheet["A3"], sheet["B3"] = "crq", "{{變更單號}}"

    layout = wb.create_sheet("LAYOUT")
    layout["A1"], layout["B1"] = "key", "value"
    rows = [
        ("header_enabled", True),
        ("header_name", "Header"),
        ("auto_fit_columns", False),
        ("image_max_width_px", 300),
        ("image_max_height_px", 200),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        layout.cell(row=i, column=1, value=k)
        layout.cell(row=i, column=2, value=v)
    layout.column_dimensions["A"].width = 28
    layout.column_dimensions["B"].width = 40
    layout.sheet_state = "hidden"

    output = EXCEL_DIR / "with_layout_template.xlsx"
    wb.save(output)
    return output


def main():
    print("開始建立 Excel 樣板範例（v2.2.1）...\n")

    sample = create_sample_template()
    print(f"[OK] 建立 {sample}")

    with_lists = create_with_lists_template()
    print(f"[OK] 建立 {with_lists}")

    with_layout = create_with_layout_template()
    print(f"[OK] 建立 {with_layout}")

    print("\n所有 Excel 樣板建立完成！")
    print("\n樣板位置：")
    print(f"  - {sample}        （基本資訊 + {{var}} 替換示範）")
    print(f"  - {with_lists}  （含 {{% for %}} 巢狀 children 展開）")
    print(f"  - {with_layout}  （含 LAYOUT metadata 範例）")


if __name__ == "__main__":
    main()
